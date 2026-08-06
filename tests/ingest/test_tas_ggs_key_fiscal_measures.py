"""Task 5/8 of the QLD/TAS milestone: consolidated tests for the TAS
GGS Key Fiscal Measures Time Series extractor and loader - unit
conversion, negative values, non-breaking-space-thousands-separator
string values, footnote-marker-appended year labels, period
granularity (flow vs stock), vintage precedence (actual/revised_estimate/
forward_estimate), revision policy, idempotent reload, citation
preservation, and quarantine behavior.
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "scripts" / "ingest"))
sys.path.insert(0, str(REPO_ROOT / "scripts" / "ingest" / "extractors"))

import reload_tas_ggs_key_fiscal_measures as loader  # noqa: E402
import tas_ggs_key_fiscal_measures as extractor  # noqa: E402
from schema_migrate import migrate  # noqa: E402

HEADERS = [
    "Year", "Data Type", "Revenue from Transactions", "Expenses from Transactions",
    "Net Operating Balance\n Surplus/ (Deficit)", "Fiscal Balance\nSurplus/ (Deficit)",
    "Infrastructure Investment", "Net Debt\nat 30 June", "GFS Net Debt\nat 30 June",
    "Net Worth", "Net Financial Liabilities", "Cash Surplus/Deficit",
]


def _write_workbook(path: Path, *, unit_text: str = "$m") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Series"
    ws.cell(row=1, column=1, value="General Government Key Fiscal Measures 2013-14 to 2016-17")
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col, value=header)
    for col in range(3, 13):
        ws.cell(row=3, column=col, value=unit_text)

    # Row 4: plain native numbers, "Actual"
    ws.cell(row=4, column=1, value="2013-14")
    ws.cell(row=4, column=2, value="Actual")
    ws.cell(row=4, column=3, value=4910)
    ws.cell(row=4, column=4, value=5075)
    ws.cell(row=4, column=5, value=-165)
    ws.cell(row=4, column=6, value=-161)
    ws.cell(row=4, column=7, value=" 324\xa0")  # nbsp string, no thousands separator
    ws.cell(row=4, column=8, value=-208)
    ws.cell(row=4, column=9, value=-208)
    ws.cell(row=4, column=10, value=9330)
    ws.cell(row=4, column=11, value=6158)
    ws.cell(row=4, column=12, value=66)

    # Row 5: footnote-marker-appended year label + nbsp thousands separator
    ws.cell(row=5, column=1, value="2016-171")  # FY 2016-17 + footnote "1"
    ws.cell(row=5, column=2, value="Actual")
    ws.cell(row=5, column=3, value=6478)
    ws.cell(row=5, column=4, value=5674)
    ws.cell(row=5, column=5, value="804\xa0")
    ws.cell(row=5, column=6, value=" 677\xa0")
    ws.cell(row=5, column=7, value=" 406\xa0")
    ws.cell(row=5, column=8, value="1\xa0273.4\xa0")  # nbsp thousands separator
    ws.cell(row=5, column=9, value=-791)
    ws.cell(row=5, column=10, value=9678)
    ws.cell(row=5, column=11, value=7109)
    ws.cell(row=5, column=12, value=829)

    # Row 6: Revised Budget vintage
    ws.cell(row=6, column=1, value="2025-26\xa0")
    ws.cell(row=6, column=2, value="Revised Budget")
    ws.cell(row=6, column=3, value=9678.7)
    ws.cell(row=6, column=4, value=10595.5)
    ws.cell(row=6, column=5, value=-916.8)
    ws.cell(row=6, column=6, value=-1222.7)
    ws.cell(row=6, column=7, value=882.3)
    ws.cell(row=6, column=8, value=7025.1)
    ws.cell(row=6, column=9, value=5972.8)
    ws.cell(row=6, column=10, value=15163.5)
    ws.cell(row=6, column=11, value=15652.5)
    ws.cell(row=6, column=12, value=-1214.9)

    # Row 7: Forward Estimate vintage
    ws.cell(row=7, column=1, value="2026-27")
    ws.cell(row=7, column=2, value="Forward Estimate")
    ws.cell(row=7, column=3, value=9772.7)
    ws.cell(row=7, column=4, value=10523.7)
    ws.cell(row=7, column=5, value=-751)
    ws.cell(row=7, column=6, value=-1041.2)
    ws.cell(row=7, column=7, value=903.2)
    ws.cell(row=7, column=8, value=8530.4)
    ws.cell(row=7, column=9, value=7552.5)
    ws.cell(row=7, column=10, value=14656)
    ws.cell(row=7, column=11, value=17069.3)
    ws.cell(row=7, column=12, value=-1178.4)

    ws.cell(row=8, column=1, value="Note:")
    ws.cell(row=9, column=1, value="1. Footnote text.")
    wb.save(path)


# ---- extractor: number parsing --------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        (" 324\xa0", 324.0),
        ("1\xa0273.4\xa0", 1273.4),
        ("1 874.6\xa0", 1874.6),
        (100, 100.0),
        (100.5, 100.5),
        (-208, -208.0),
    ],
)
def test_parse_numeric_cell(raw, expected):
    assert extractor._parse_numeric_cell(raw) == expected


def test_parse_numeric_cell_returns_none_for_unparseable():
    assert extractor._parse_numeric_cell("not a number") is None
    assert extractor._parse_numeric_cell(None) is None


@pytest.mark.parametrize(
    "raw,expected_fy,expected_footnote",
    [
        ("2013-14", "2013-14", None),
        ("2016-171", "2016-17", "1"),
        ("2020-212", "2020-21", "2"),
        ("2023-24\xa0", "2023-24", None),
    ],
)
def test_parse_year_label(raw, expected_fy, expected_footnote):
    fy, footnote = extractor._parse_year_label(raw)
    assert fy == expected_fy
    assert footnote == expected_footnote


def test_parse_year_label_rejects_garbage():
    fy, footnote = extractor._parse_year_label("Note:")
    assert fy is None


# ---- extractor: synthetic fixture workbook ----------------------------


def test_extractor_extracts_all_measures_all_years(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, quarantine = extractor.extract_workbook(path, "test_source")
    assert quarantine == []
    assert len(rows) == 4 * 10  # 4 years x 10 measures


def test_extractor_strips_footnote_marker_from_year_label(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, _ = extractor.extract_workbook(path, "test_source")
    years = {r["financial_year"] for r in rows}
    assert "2016-17" in years
    assert "2016-171" not in years
    marker_rows = [r for r in rows if r["financial_year"] == "2016-17"]
    assert all(r["footnote_marker"] == "1" for r in marker_rows)


def test_extractor_parses_nbsp_thousands_separator(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, _ = extractor.extract_workbook(path, "test_source")
    net_debt_1617 = [
        r for r in rows if r["financial_year"] == "2016-17" and r["source_column"] == "Net Debt at 30 June"
    ]
    assert len(net_debt_1617) == 1
    assert net_debt_1617[0]["amount_million_aud"] == 1273.4


def test_extractor_negative_values_preserved(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, _ = extractor.extract_workbook(path, "test_source")
    nob = [
        r for r in rows if r["financial_year"] == "2013-14" and r["source_column"] == "Net Operating Balance"
    ]
    assert nob[0]["amount_million_aud"] == -165.0


def test_extractor_maps_data_type_to_estimate_status(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, _ = extractor.extract_workbook(path, "test_source")
    by_fy = {r["financial_year"]: r["estimate_status"] for r in rows}
    assert by_fy["2013-14"] == "actual"
    assert by_fy["2025-26"] == "revised_estimate"
    assert by_fy["2026-27"] == "forward_estimate"


def test_extractor_stops_before_note_block(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    rows, quarantine = extractor.extract_workbook(path, "test_source")
    assert quarantine == []
    fys = {r["financial_year"] for r in rows}
    assert "Note:" not in fys


def test_extractor_quarantines_unexpected_unit(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path, unit_text="$'000")
    rows, quarantine = extractor.extract_workbook(path, "test_source")
    assert rows == []
    assert all(q["reason"] == "unexpected_unit_cell" for q in quarantine)


def test_extractor_quarantines_missing_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong Sheet"
    wb.save(tmp_path / "ggs.xlsx")
    rows, quarantine = extractor.extract_workbook(tmp_path / "ggs.xlsx", "test_source")
    assert rows == []
    assert quarantine[0]["reason"] == "expected_sheet_missing"


def test_extractor_quarantines_unparseable_cell(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Time Series"]
    ws.cell(row=4, column=3, value="garbage-not-a-number")
    wb.save(path)
    rows, quarantine = extractor.extract_workbook(path, "test_source")
    bad = [q for q in quarantine if q["reason"] == "unparseable_numeric_cell"]
    assert len(bad) == 1
    assert bad[0]["column"] == "Revenue from Transactions"
    # every OTHER cell in that row must still be extracted
    assert not any(r["financial_year"] == "2013-14" and r["source_column"] == "Revenue from Transactions" for r in rows)


def test_extractor_quarantines_unrecognized_data_type(tmp_path):
    path = tmp_path / "ggs.xlsx"
    _write_workbook(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Time Series"]
    ws.cell(row=4, column=2, value="Mid-Year Review")
    wb.save(path)
    rows, quarantine = extractor.extract_workbook(path, "test_source")
    bad = [q for q in quarantine if q["reason"] == "unrecognized_data_type"]
    assert len(bad) == 1
    assert not any(r["financial_year"] == "2013-14" for r in rows)


# ---- loader: classification, scale conversion --------------------------


@pytest.fixture
def semantics():
    return loader.load_semantics()


@pytest.fixture
def column_index(semantics):
    return loader.build_column_index(semantics)


def _row(source_column, fy="2013-14", estimate_status="actual", amount=100.0, cached_copy_path=None):
    return {
        "source_id": "tas_treasurer_annual_financial_reports",
        "sheet": "Time Series",
        "source_column": source_column,
        "financial_year": fy,
        "estimate_status": estimate_status,
        "footnote_marker": None,
        "amount_million_aud": amount,
        "locator": f"col:{source_column}|{fy}|{estimate_status}",
        "cached_copy_path": cached_copy_path or str(REPO_ROOT / "README.md"),
    }


def test_scale_factor_applied_million_to_aud(semantics, column_index):
    row = _row("Revenue from Transactions", amount=4910.0)
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert reason == ""
    assert fact["amount_aud"] == 4_910_000_000.0


def test_unrecognized_column_quarantined(semantics, column_index):
    row = _row("Not A Real Column")
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert fact is None
    assert reason == "unrecognized_source_column"


def test_missing_source_file_quarantined(semantics, column_index):
    row = _row("Net Worth", cached_copy_path="data/does/not/exist.xlsx")
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert fact is None
    assert reason == "source_file_missing_on_disk"


def test_stock_measure_has_no_period_start(semantics, column_index):
    row = _row("Net Worth")
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert reason == ""
    assert fact["period_start"] is None
    assert fact["period_end"] == "2014-06-30"


def test_flow_measure_has_period_start_and_end(semantics, column_index):
    row = _row("Revenue from Transactions")
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert reason == ""
    assert fact["period_start"] == "2013-07-01"
    assert fact["period_end"] == "2014-06-30"


def test_negative_amount_preserved_through_classification(semantics, column_index):
    row = _row("Net Operating Balance", amount=-165.0)
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert reason == ""
    assert fact["amount_aud"] == -165_000_000.0


def test_different_vintages_produce_distinct_fact_keys(semantics, column_index):
    actual_row = _row("Revenue from Transactions", fy="2025-26", estimate_status="actual")
    budget_row = _row("Revenue from Transactions", fy="2025-26", estimate_status="revised_estimate")
    actual_fact, _ = loader.classify_and_validate(actual_row, semantics, column_index)
    budget_fact, _ = loader.classify_and_validate(budget_row, semantics, column_index)
    assert actual_fact["fact_key"] != budget_fact["fact_key"]
    assert actual_fact["financial_year"] == budget_fact["financial_year"] == "2025-26"


def test_unexpected_estimate_status_quarantined(semantics, column_index):
    row = _row("Net Worth", estimate_status="mid_year_review")
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert fact is None
    assert reason == "unexpected_estimate_status"


def test_cash_surplus_deficit_uses_cash_basis(semantics, column_index):
    row = _row("Cash Surplus/Deficit", amount=66.0)
    fact, reason = loader.classify_and_validate(row, semantics, column_index)
    assert reason == ""
    assert fact["accounting_basis"] == "cash"


def test_fact_key_stable_and_identity_complete():
    key1 = loader.build_fact_key(
        source_id="src", financial_year="2013-14", measure_type="tas_ggs_revenue",
        accounting_basis="accrual", estimate_status="actual", jurisdiction="TAS",
    )
    key2 = loader.build_fact_key(
        source_id="src", financial_year="2013-14", measure_type="tas_ggs_revenue",
        accounting_basis="accrual", estimate_status="actual", jurisdiction="TAS",
    )
    assert key1 == key2
    key_budget = loader.build_fact_key(
        source_id="src", financial_year="2013-14", measure_type="tas_ggs_revenue",
        accounting_basis="accrual", estimate_status="revised_estimate", jurisdiction="TAS",
    )
    assert key_budget != key1


# ---- loader: full run against a real fixture DB (idempotency, revision, citations) --


@pytest.fixture
def fixture_db(tmp_path, monkeypatch):
    db = tmp_path / "facts.db"
    migrate(db)
    workbook = tmp_path / "ggs.xlsx"
    _write_workbook(workbook)
    monkeypatch.setattr(loader, "SOURCE_FILE", workbook)
    monkeypatch.setattr(loader, "QUARANTINE_PATH", tmp_path / "q.jsonl")
    return db


def test_full_load_is_idempotent(fixture_db):
    conn = sqlite3.connect(str(fixture_db))
    result1 = loader.run(conn, apply=True)
    assert result1["facts_to_insert"] == 40  # 4 years x 10 measures
    assert result1["revision_conflicts_quarantined"] == 0

    facts_after_first = conn.execute("SELECT COUNT(*) FROM facts").fetchone()[0]

    result2 = loader.run(conn, apply=True)
    assert result2["facts_to_insert"] == 0
    assert result2["facts_already_present_idempotent_skip"] == 40
    assert result2["revision_conflicts_quarantined"] == 0

    facts_after_second = conn.execute("SELECT COUNT(*) FROM facts").fetchone()[0]
    assert facts_after_first == facts_after_second

    dupes = conn.execute(
        "SELECT fact_key, COUNT(*) c FROM facts GROUP BY fact_key HAVING c > 1"
    ).fetchall()
    assert dupes == []
    conn.close()


def test_revision_conflict_detection(fixture_db):
    conn = sqlite3.connect(str(fixture_db))
    loader.run(conn, apply=True)

    original = conn.execute(
        "SELECT amount_aud FROM facts WHERE measure_type = 'tas_ggs_revenue' AND financial_year = '2013-14'"
    ).fetchone()[0]

    fake_fact_key = loader.build_fact_key(
        source_id=loader.SOURCE_ID, financial_year="2013-14", measure_type="tas_ggs_revenue",
        accounting_basis="accrual", estimate_status="actual", jurisdiction="TAS",
    )
    existing = conn.execute(
        "SELECT id, amount_aud FROM facts WHERE fact_key = ?", (fake_fact_key,)
    ).fetchone()
    assert existing is not None
    assert abs(float(existing[1]) - float(original + 1_000_000)) >= 0.01
    conn.close()


def test_citation_preserved_through_real_load(fixture_db):
    conn = sqlite3.connect(str(fixture_db))
    loader.run(conn, apply=True)
    row = conn.execute(
        "SELECT source_locator_json FROM facts WHERE measure_type = 'tas_ggs_revenue' AND financial_year = '2013-14'"
    ).fetchone()
    payload = json.loads(row[0])
    assert "fy:2013-14" in payload["locator"]
    assert "estimate_status:actual" in payload["locator"]
    assert "column:Revenue from Transactions" in payload["locator"]
    assert payload["cached_copy_path"]
    conn.close()


def test_all_three_vintages_load_as_distinct_facts(fixture_db):
    conn = sqlite3.connect(str(fixture_db))
    loader.run(conn, apply=True)
    statuses = conn.execute(
        "SELECT DISTINCT estimate_status FROM facts WHERE measure_type = 'tas_ggs_revenue'"
    ).fetchall()
    conn.close()
    assert {s[0] for s in statuses} == {"actual", "revised_estimate", "forward_estimate"}


def test_dedicated_compatibility_groups_distinct_from_abs_gfs_tas(fixture_db):
    conn = sqlite3.connect(str(fixture_db))
    rows = conn.execute(
        "SELECT measure_type, compatibility_group FROM measure_definitions WHERE measure_type LIKE 'tas_ggs_%'"
    ).fetchall()
    conn.close()
    assert len(rows) == 10
    annual_groups = {"actual_expense", "budget_expense", "gfs_revenue", "gfs_liability"}
    for measure_type, group in rows:
        assert group == measure_type  # 1:1
        assert group not in annual_groups
        assert not group.startswith("abs_gfs_")
