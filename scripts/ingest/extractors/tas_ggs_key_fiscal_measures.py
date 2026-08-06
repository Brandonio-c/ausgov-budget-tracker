#!/usr/bin/env python3
"""Extractor for the Tasmanian Department of Treasury and Finance's
"GGS Key Fiscal Measures Time Series" workbook - see
ops/reports/qld-tas-family-inventory-20260806T171537Z.md for the full
cell-level inventory this is built from.

Genuinely different shape from every row-labeled statement extractor
in this repo (vic_bpo.py, vic_afs.py, vic_bpo_soce_admin.py): this is a
wide time series - one row per financial year, one column per measure,
a `Data Type` column giving that year's vintage (Actual / Revised
Budget / Forward Estimate). Two real data-quality issues, confirmed by
reading the workbook with `openpyxl` (`data_only=True`) rather than
trusting pandas's type-inferred view:

1. Measure cells are inconsistently typed - some are native Excel
   numbers, others are strings with a leading space and a trailing
   non-breaking space (`\\xa0`), and some larger values use an embedded
   non-breaking space (or, in one cell, a literal space) as a
   thousands separator inside the string (e.g. `'1\\xa0273.4\\xa0'`).
2. Two year labels have a footnote-reference digit appended directly
   to the `YYYY-YY` string with no separator (`'2016-171'`,
   `'2020-212'`); several others carry a harmless trailing
   non-breaking space.

This module only extracts and quarantines; it does not write to
facts.db. See config/measure-semantics/tas_ggs_key_fiscal_measures.yaml
for which source columns map to which measure_type, and
scripts/ingest/reload_tas_ggs_key_fiscal_measures.py for classification/
validation/loading.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[3]
QUARANTINE_DIR = REPO_ROOT / "data" / "staging" / "quarantine"

EXPECTED_SHEET = "Time Series"
EXPECTED_UNIT = "$m"
HEADER_ROW = 2
UNIT_ROW = 3
FIRST_DATA_ROW = 4
YEAR_COL = 1
DATA_TYPE_COL = 2

YEAR_LABEL_RE = re.compile(r"^(\d{4}-\d{2})(\d*)\s*$")
DATA_TYPE_MAP = {
    "Actual": "actual",
    "Revised Budget": "revised_estimate",
    "Forward Estimate": "forward_estimate",
}

MEASURE_COLUMNS: dict[int, str] = {
    3: "Revenue from Transactions",
    4: "Expenses from Transactions",
    5: "Net Operating Balance",
    6: "Fiscal Balance",
    7: "Infrastructure Investment",
    8: "Net Debt at 30 June",
    9: "GFS Net Debt at 30 June",
    10: "Net Worth",
    11: "Net Financial Liabilities",
    12: "Cash Surplus/Deficit",
}


def _relative_or_str(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def _parse_year_label(raw: str) -> tuple[str | None, str | None]:
    """Returns (financial_year, footnote_marker_or_None). None for
    financial_year if the label doesn't match the expected shape."""
    if not isinstance(raw, str):
        return None, None
    cleaned = raw.replace("\xa0", " ").strip()
    m = YEAR_LABEL_RE.match(cleaned)
    if not m:
        return None, None
    return m.group(1), (m.group(2) or None)


def _parse_numeric_cell(value) -> float | None:
    """Returns a float, or None if the cell cannot be confidently
    parsed as a number - callers must quarantine on None, never guess."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\xa0", " ")
        cleaned = re.sub(r"\s+", "", cleaned)
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def extract_workbook(path: Path, source_id: str) -> tuple[list[dict], list[dict]]:
    rows: list[dict] = []
    quarantine: list[dict] = []

    wb = openpyxl.load_workbook(path, data_only=True)
    if EXPECTED_SHEET not in wb.sheetnames:
        return rows, [{"reason": "expected_sheet_missing", "sheet": EXPECTED_SHEET}]
    ws = wb[EXPECTED_SHEET]

    unit_cells = {
        col: ws.cell(row=UNIT_ROW, column=col).value for col in MEASURE_COLUMNS
    }
    bad_units = {col: v for col, v in unit_cells.items() if v != EXPECTED_UNIT}
    if bad_units:
        return rows, [
            {
                "reason": "unexpected_unit_cell",
                "sheet": EXPECTED_SHEET,
                "column": col,
                "unit_cell": repr(v),
            }
            for col, v in bad_units.items()
        ]

    row_idx = FIRST_DATA_ROW
    while True:
        year_raw = ws.cell(row=row_idx, column=YEAR_COL).value
        if year_raw is None or (isinstance(year_raw, str) and year_raw.strip().lower() == "note:"):
            break
        if not isinstance(year_raw, str) or not year_raw.strip():
            row_idx += 1
            continue

        financial_year, footnote_marker = _parse_year_label(year_raw)
        if financial_year is None:
            quarantine.append(
                {
                    "reason": "unparsed_year_label",
                    "sheet": EXPECTED_SHEET,
                    "row": row_idx,
                    "raw_value": repr(year_raw),
                }
            )
            row_idx += 1
            continue

        data_type_raw = ws.cell(row=row_idx, column=DATA_TYPE_COL).value
        data_type_clean = data_type_raw.strip() if isinstance(data_type_raw, str) else None
        estimate_status = DATA_TYPE_MAP.get(data_type_clean) if data_type_clean else None
        if estimate_status is None:
            quarantine.append(
                {
                    "reason": "unrecognized_data_type",
                    "sheet": EXPECTED_SHEET,
                    "row": row_idx,
                    "financial_year": financial_year,
                    "raw_value": repr(data_type_raw),
                }
            )
            row_idx += 1
            continue

        for col, source_column in MEASURE_COLUMNS.items():
            cell = ws.cell(row=row_idx, column=col)
            value = _parse_numeric_cell(cell.value)
            if value is None:
                quarantine.append(
                    {
                        "reason": "unparseable_numeric_cell",
                        "sheet": EXPECTED_SHEET,
                        "row": row_idx,
                        "column": source_column,
                        "financial_year": financial_year,
                        "raw_value": repr(cell.value),
                    }
                )
                continue
            rows.append(
                {
                    "source_id": source_id,
                    "sheet": EXPECTED_SHEET,
                    "source_column": source_column,
                    "financial_year": financial_year,
                    "estimate_status": estimate_status,
                    "footnote_marker": footnote_marker,
                    "amount_million_aud": value,
                    "locator": (
                        f"source_id:{source_id} | sheet:{EXPECTED_SHEET} | "
                        f"cell:{cell.coordinate} | fy:{financial_year} | "
                        f"estimate_status:{estimate_status} | column:{source_column}"
                    ),
                    "cached_copy_path": _relative_or_str(path),
                }
            )
        row_idx += 1

    return rows, quarantine


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--source-id", default="tas_treasurer_annual_financial_reports")
    parser.add_argument(
        "--path",
        type=Path,
        default=REPO_ROOT
        / "data/raw/state/tas_treasurer_annual_financial_reports/snapshots/20260724T170239Z/files/GGS-Key-Fiscal-Measures-Time-Series.xlsx",
    )
    args = parser.parse_args()

    rows, quarantine = extract_workbook(args.path, args.source_id)
    QUARANTINE_DIR.mkdir(parents=True, exist_ok=True)
    with open(QUARANTINE_DIR / "tas_ggs_key_fiscal_measures_extract_quarantine.jsonl", "w", encoding="utf-8") as fh:
        for q in quarantine:
            fh.write(json.dumps(q) + "\n")

    print(json.dumps({"rows_extracted": len(rows), "rows_quarantined": len(quarantine)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
