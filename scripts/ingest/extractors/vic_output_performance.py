#!/usr/bin/env python3
"""Extract the seven dollar-denominated VIC DTF output-cost rows."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
SOURCE_ID = "vic_output_performance_measures_2024_25"
SOURCE_FILE = REPO_ROOT / "data/raw/state/vic_output_performance_measures_2024_25/snapshots/20260724T190604Z/files/Output-performance-measures-2024-25.xlsx"
EXPECTED_SHEETS = [
    "Budget and Financial Advice",
    "Revenue Management and Adm Serv",
    "Economic and Policy Advice",
    "Economic Regulatory Services",
    "Commercial and Infra Advice",
    "Infrastructure Victoria",
    "Industrial Relations",
]


def _number(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(" ", "").replace(",", ""))
        except ValueError:
            return None
    return None


def extract_workbook(path: Path = SOURCE_FILE) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    quarantine: list[dict] = []
    try:
        cached = str(path.resolve().relative_to(REPO_ROOT))
    except ValueError:
        cached = str(path)
    for sheet in EXPECTED_SHEETS:
        if sheet not in workbook.sheetnames:
            quarantine.append({"reason": "missing_expected_output_sheet", "sheet": sheet})
            continue
        worksheet = workbook[sheet]
        matches = []
        for row_number, values in enumerate(worksheet.iter_rows(values_only=True), 1):
            label, unit, actual, target = (list(values[:4]) + [None] * 4)[:4]
            if label == "Total output cost":
                matches.append((row_number, unit, actual, target))
        if len(matches) != 1:
            quarantine.append({"reason": "missing_total_output_cost_row", "sheet": sheet, "match_count": len(matches)})
            continue
        row_number, unit, actual_raw, target_raw = matches[0]
        if unit != "$ million":
            quarantine.append({"reason": "unit_not_aud_million", "sheet": sheet, "row": row_number, "unit": unit})
            continue
        actual, target = _number(actual_raw), _number(target_raw)
        if actual is None or target is None:
            quarantine.append({"reason": "non_numeric_actual_or_target", "sheet": sheet, "row": row_number})
            continue
        for status, amount, column in (("actual", actual, "C"), ("budget", target, "D")):
            rows.append({
                "source_id": SOURCE_ID,
                "financial_year": "2024-25",
                "publication_date": "2025-10-01",
                "output_name": sheet,
                "row_label": "Total output cost",
                "measure_type": "vic_output_total_cost",
                "estimate_status": status,
                "amount_million_aud": amount,
                "column_header_original": "2024-25 actual" if status == "actual" else "2024-25 target",
                "locator": f"source_id:{SOURCE_ID} | workbook:{path.name} | sheet:{sheet} | cell:{column}{row_number} | row:Total output cost | fy:2024-25 | estimate_status:{status}",
                "cached_copy_path": cached,
            })
    workbook.close()
    return rows, quarantine
