#!/usr/bin/env python3
"""Hang DSS recipient demographics under Social protection PBS parents (FY 2024-25).

Sources (already on disk):
  - JobSeeker monthly profile (State + Age) — June 2025 snapshot
  - Payment demographics quarterly (State × payment) — March 2025

Counts are stored as amount_aud for sunburst sizing; they must NOT be treated as
GFS dollars (compatibility_group=count / measure_type=recipient_count).
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[3]
OUT_CSV = REPO_ROOT / "data/staging/breakdowns/dss_payment_demographics_2024_25.csv"
FY = "2024-25"
LANDING = "https://data.gov.au/data/dataset/dss-payment-demographic-data"

JOBSEEKER_XLSX = (
    REPO_ROOT
    / "data/raw/federal/dss_jobseeker_monthly_profile/snapshots/20260721T024615Z/files"
    / "expanded-dss-jobseeker-payment-and-youth-allowance-recipients-monthly-profile-june-2025.xlsx"
)
DEMO_XLSX = (
    REPO_ROOT
    / "data/raw/federal/dss_payment_demographics_quarterly_related_view"
    / "snapshots/20260724T192110Z/files/dss-demographics-march-2025-final.xlsx"
)

# Payment column / sheet → PBS cascade parent path
PAYMENT_PARENTS = {
    "JobSeeker Payment": (
        "Social security and welfare / Assistance to the unemployed and the sick / "
        "Job Seeker Income Support"
    ),
    "Age Pension": (
        "Social security and welfare / Assistance to the aged / Support for Seniors"
    ),
    "Disability Support Pension": (
        "Social security and welfare / Assistance to people with disabilities / "
        "Financial Support for People with Disability"
    ),
    "Carer Payment": (
        "Social security and welfare / Assistance to people with disabilities / "
        "Financial Support for Carers"
    ),
    "Parenting Payment Single": (
        "Social security and welfare / Assistance to families with children / "
        "Support for Families"
    ),
    "Parenting Payment Partnered": (
        "Social security and welfare / Assistance to families with children / "
        "Support for Families"
    ),
}

JOBSEEKER_PARENT = PAYMENT_PARENTS["JobSeeker Payment"]
STATES = {
    "New South Wales",
    "Victoria",
    "Queensland",
    "South Australia",
    "Western Australia",
    "Tasmania",
    "Northern Territory",
    "Australian Capital Territory",
}


def _clean(s: str, limit: int = 80) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return (s.replace(" / ", " – ")[:limit] or "Unknown").strip()


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_jobseeker(rows: list[dict]) -> None:
    import openpyxl

    if not JOBSEEKER_XLSX.is_file():
        return
    wb = openpyxl.load_workbook(JOBSEEKER_XLSX, read_only=True, data_only=True)
    resource = JOBSEEKER_XLSX.as_posix()

    # State
    ws = wb["Table 3 - By State"]
    for row in ws.iter_rows(values_only=True):
        label = row[0]
        total = _num(row[3]) if len(row) > 3 else None
        if not isinstance(label, str) or label not in STATES or total is None or total <= 0:
            continue
        parent = JOBSEEKER_PARENT
        cat = f"{parent} / Recipients by state / {_clean(label)}"
        rows.append(
            {
                "fy": FY,
                "amount": f"{total:.0f}",
                "category": cat,
                "estimate_status": "actual",
                "locator": "dss:jobseeker:june-2025:by-state",
                "landing_url": LANDING,
                "resource_url": resource,
            }
        )

    # Age
    ws = wb["Table 4 - By Age and Gender"]
    for row in ws.iter_rows(values_only=True):
        label = row[0]
        total = _num(row[3]) if len(row) > 3 else None
        if not isinstance(label, str) or total is None or total <= 0:
            continue
        if label in ("Total Recipients", "By Age Group") or "Data" in label:
            continue
        if "year" not in label.lower() and "over" not in label.lower():
            continue
        cat = f"{JOBSEEKER_PARENT} / Recipients by age / {_clean(label)}"
        rows.append(
            {
                "fy": FY,
                "amount": f"{total:.0f}",
                "category": cat,
                "estimate_status": "actual",
                "locator": "dss:jobseeker:june-2025:by-age",
                "landing_url": LANDING,
                "resource_url": resource,
            }
        )
    wb.close()


def _extract_demo_state(rows: list[dict]) -> None:
    import openpyxl

    if not DEMO_XLSX.is_file():
        return
    wb = openpyxl.load_workbook(DEMO_XLSX, read_only=True, data_only=True)
    ws = wb["State"]
    resource = DEMO_XLSX.as_posix()
    header = None
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if vals and vals[0] == "State":
            header = vals
            break
    if not header:
        wb.close()
        return
    col_idx = {h: i for i, h in enumerate(header) if h}
    for row in ws.iter_rows(values_only=True):
        if not row or not isinstance(row[0], str):
            continue
        state = row[0].strip()
        if state not in STATES:
            continue
        for payment, parent in PAYMENT_PARENTS.items():
            if payment == "JobSeeker Payment":
                continue  # covered by jobseeker profile
            idx = col_idx.get(payment)
            if idx is None:
                continue
            total = _num(row[idx] if idx < len(row) else None)
            if total is None or total <= 0:
                continue
            cat = f"{parent} / Recipients by state / {_clean(state)}"
            rows.append(
                {
                    "fy": FY,
                    "amount": f"{total:.0f}",
                    "category": cat,
                    "estimate_status": "actual",
                    "locator": f"dss:demographics:march-2025:state:{payment}",
                    "landing_url": LANDING,
                    "resource_url": resource,
                }
            )
    wb.close()


def extract() -> list[dict]:
    rows: list[dict] = []
    _extract_jobseeker(rows)
    _extract_demo_state(rows)
    # Dedupe leaves
    best: dict[str, dict] = {}
    for r in rows:
        best[r["category"]] = r
    rows = list(best.values())

    # Emit folder aggregates so same_group cascade can nest (needs a fact per node)
    folder_sums: dict[str, float] = {}
    for r in rows:
        parts = r["category"].split(" / ")
        if len(parts) < 2:
            continue
        # parent path = everything except last segment
        parent = " / ".join(parts[:-1])
        folder_sums[parent] = folder_sums.get(parent, 0.0) + float(r["amount"])
    for folder, total in folder_sums.items():
        if folder in best:
            continue
        rows.append(
            {
                "fy": FY,
                "amount": f"{total:.0f}",
                "category": folder,
                "estimate_status": "actual",
                "locator": "dss:demographics:folder-aggregate",
                "landing_url": LANDING,
                "resource_url": str(JOBSEEKER_XLSX),
            }
        )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "fy",
        "amount",
        "category",
        "estimate_status",
        "locator",
        "landing_url",
        "resource_url",
    ]
    with OUT_CSV.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print({"wrote": str(OUT_CSV), "rows": len(rows), "folders": len(folder_sums)})
    return rows


def main() -> int:
    extract()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
