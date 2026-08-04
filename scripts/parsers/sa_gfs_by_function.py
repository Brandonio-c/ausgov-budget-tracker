"""Parser for the SA Government Finance Statistics 'GPC by ETF' workbook.

One sheet per financial year (e.g. "2015-16"). Each sheet is a matrix: rows
are GPC (Government Purpose Classification - i.e. function, e.g. "Police
services"), columns are ETF (Economic Type of Flow, e.g. "Wages, sal & supp")
plus a trailing 'Total GPC' column. Figures are in $'000 (SA GFS convention).

The 'Total GPC' column is a row aggregate and is excluded to avoid
double-counting against its own components.
"""
import openpyxl
from openpyxl.utils import get_column_letter

from .common import SpendingRow


def parse(raw_file, meta: dict) -> list[SpendingRow]:
    wb = openpyxl.load_workbook(raw_file, read_only=True, data_only=True)
    rows: list[SpendingRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if len(sheet_rows) < 5:
            continue

        etf_header = sheet_rows[2]  # codes row, last cell literal 'Total GPC'
        description_row = sheet_rows[3]  # e.g. ['GPC', 'Description', 'Funded super', ...]

        total_col_idx = None
        for i, v in enumerate(etf_header):
            if isinstance(v, str) and v.strip().lower() == "total gpc":
                total_col_idx = i
                break

        subcategory_cols = [
            i
            for i in range(2, len(description_row))
            if i != total_col_idx and isinstance(description_row[i], str) and description_row[i].strip()
        ]

        for source_row_index, row in enumerate(sheet_rows[4:], start=5):
            gpc_description = row[1]
            if not isinstance(gpc_description, str) or not gpc_description.strip():
                continue

            for i in subcategory_cols:
                value = row[i] if i < len(row) else None
                if not isinstance(value, (int, float)) or value == 0:
                    continue
                context_columns = [1, *range(max(2, i - 1), min(len(description_row), i + 2))]
                context_start = max(5, source_row_index - 1)
                context_end = min(len(sheet_rows), source_row_index + 1)
                context_rows = [
                    [sheet_rows[row_number - 1][column] if column < len(sheet_rows[row_number - 1]) else None for column in context_columns]
                    for row_number in range(context_start, context_end + 1)
                ]
                selected_cells = f"B{context_start}:B{context_end}, {get_column_letter(context_columns[1] + 1)}{context_start}:{get_column_letter(context_columns[-1] + 1)}{context_end}"
                rows.append(
                    SpendingRow(
                        financial_year=sheet_name,
                        level_of_government="state",
                        jurisdiction="SA",
                        category=gpc_description.strip(),
                        subcategory=description_row[i].strip(),
                        department=None,
                        amount_aud=round(value * 1_000, 2),
                        source_document_name=f"{meta['source_document_name']} — {sheet_name}",
                        source_url=meta["source_url"],
                        retrieved_at=meta["retrieved_at"],
                        source_context={
                            "source_type": "spreadsheet",
                            "sheet_name": sheet_name,
                            "cell_range": selected_cells,
                            "columns": ["GPC description", *[str(description_row[column]).strip() for column in context_columns[1:]]],
                            "rows": context_rows,
                            "highlight": {
                                "row_index": source_row_index - context_start,
                                "column_index": context_columns.index(i),
                                "cell": f"{get_column_letter(i + 1)}{source_row_index}",
                            },
                            "unit": "AUD thousands",
                            "note": "The highlighted workbook value is multiplied by 1,000 for the normalized AUD amount.",
                        },
                    )
                )
    return rows
