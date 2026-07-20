"""Parser for the federal 'Note 3 Function Statement' workbook.

One sheet per financial year (e.g. "2024-25"). Each sheet has a header block
followed by expense-by-function rows with cumulative year-to-date monthly
columns in $m (July .. YTD May - this monthly series never includes a June
column; June is reconciled separately in the annual Final Budget Outcome).
We take the last non-null YTD column as the best-available actual for the
year and label it accordingly rather than presenting it as a final audited
total.

'Total expenses' is an aggregate row and is excluded to avoid double-counting
against its own components.
"""
import openpyxl
from openpyxl.utils import get_column_letter

from .common import SpendingRow

EXCLUDED_ROW_LABELS = {"total expenses"}


def parse(raw_file, meta: dict) -> list[SpendingRow]:
    wb = openpyxl.load_workbook(raw_file, read_only=True, data_only=True)
    rows: list[SpendingRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if len(sheet_rows) < 3:
            continue

        data_start_row = next(
            (
                row_index
                for row_index, row in enumerate(sheet_rows, start=1)
                if isinstance(row[0], str)
                and any(isinstance(value, (int, float)) for value in row[1:])
            ),
            None,
        )
        if data_start_row is None:
            continue

        def column_heading(column: int) -> str:
            parts: list[str] = []
            for header_row in sheet_rows[: data_start_row - 1]:
                if column >= len(header_row) or not isinstance(header_row[column], str):
                    continue
                for part in header_row[column].splitlines():
                    part = part.strip()
                    if not part or part.upper() == "ACTUAL" or part.replace("-", "").isdigit():
                        continue
                    if part not in parts:
                        parts.append(part)
            return " ".join(parts) or get_column_letter(column + 1)

        for source_row_index, row in enumerate(sheet_rows[data_start_row - 1 :], start=data_start_row):
            label = row[0]
            if not isinstance(label, str):
                continue
            label = label.strip()
            if not label or label.lower() in EXCLUDED_ROW_LABELS:
                continue
            numeric_columns = [i for i, value in enumerate(row[1:], start=1) if isinstance(value, (int, float))]
            if not numeric_columns:
                continue  # section header row (e.g. "Expenses by function"), no figures

            value_column = numeric_columns[-1]
            amount_m = row[value_column]  # latest cumulative YTD actual, in $m
            context_columns = [0, *range(max(1, value_column - 2), value_column + 1)]
            context_start = max(data_start_row, source_row_index - 1)
            context_end = min(len(sheet_rows), source_row_index + 1)
            context_rows = [
                [sheet_rows[i - 1][column] if column < len(sheet_rows[i - 1]) else None for column in context_columns]
                for i in range(context_start, context_end + 1)
            ]
            context_headers = ["Expense function", *[column_heading(column) for column in context_columns[1:]]]
            selected_cells = f"A{context_start}:A{context_end}, {get_column_letter(context_columns[1] + 1)}{context_start}:{get_column_letter(value_column + 1)}{context_end}"
            rows.append(
                SpendingRow(
                    financial_year=sheet_name,
                    level_of_government="federal",
                    jurisdiction="Commonwealth",
                    category=label,
                    subcategory=None,
                    department=None,
                    amount_aud=round(amount_m * 1_000_000, 2),
                    source_document_name=f"{meta['source_document_name']} — {sheet_name} (cumulative YTD actual, latest available month)",
                    source_url=meta["source_url"],
                    retrieved_at=meta["retrieved_at"],
                    source_context={
                        "source_type": "spreadsheet",
                        "sheet_name": sheet_name,
                        "cell_range": selected_cells,
                        "columns": context_headers,
                        "rows": context_rows,
                        "highlight": {
                            "row_index": source_row_index - context_start,
                            "column_index": context_columns.index(value_column),
                            "cell": f"{get_column_letter(value_column + 1)}{source_row_index}",
                        },
                        "unit": "AUD millions",
                        "note": "The highlighted workbook value is multiplied by 1,000,000 for the normalized AUD amount.",
                    },
                )
            )
    return rows
